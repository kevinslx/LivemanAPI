import sys
import xlrd
from openpyxl.reader.excel import *
import openpyxl
import os
import time
import requests

path = os.path.abspath('.')
excelPath = path + '/Data/case.xlsx'


class ExcelData(object):
	def __init__(self, excelFile, excelPath=''):
		self.excelFile = excelFile
		self.excelPath = excelPath
		self.caseList = []
		self.resultList = [{'ID': 1, 'ActualResult': 'Beijing', 'Result': 'Pass', 'ResTime': '3'}, {'ID': 2, 'ActualResult': 'Failure', 'Result': 'Fail', 'ResTime': '6'}]

	def get_case_list(self):
		exceldata = xlrd.open_workbook(self.excelFile)
		sheet = exceldata.sheet_by_index(0)
		caseid = sheet.col_values(0, 1)
		host = sheet.col_values(2, 1)
		url = sheet.col_values(3, 1)
		method = sheet.col_values(4, 1)
		params = sheet.col_values(5, 1)
		for i in range(len(caseid)):
			temp = {}
			temp['CaseID'] = int(caseid[i])
			temp['Host'] = host[i]
			temp['URL'] = url[i]
			temp['Method'] = method[i]
			temp['Params'] = params[i]
			self.caseList.append(temp)
		print(self.caseList)
		return self.caseList

	def write_case_result(self):
		exceldata = load_workbook(self.excelFile)
		sheet = exceldata.get_sheet_by_name('Sheet1')
		for result in self.resultList:
			print(result)
			for i in range(len(self.resultList)):
				print(i)
				if result['ID'] == sheet.cell(row=i+2, column=1).value:
					print(i)
					sheet.cell(row=i+2, column=6).value = result['ActualResult']
					sheet.cell(row=i+2, column=7).value = result['Result']
					sheet.cell(row=i+2, column=8).value = result['ResTime']
		exceldata.save(self.excelFile)


class TestAPI(object):
	def __init__(self, caselist):
		self.caselist = caselist

	def test_API(self):
		for case in self.caselist:
			paramstemp = case['Params'].split(':')
			if paramstemp == ['']:
				params = {}
			else:
				params = {paramstemp[0]:paramstemp[1]}
			r = requests.request(method=case['Method'], url=case['Host']+case['URL'], params=params)
			print(r.text)


excel = ExcelData(excelPath)
caselist = excel.get_case_list()
#excel.write_case_result()

test = TestAPI(caselist)
test.test_API()