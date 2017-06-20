import sys
import xlrd
from openpyxl.reader.excel import *
import openpyxl
import os
import time
import requests

path = os.path.abspath('.')
excelPath = path + '/Data/weather.xlsx'


class ExcelData(object):
	def __init__(self, excelFile, excelPath=''):
		self.excelFile = excelFile
		self.excelPath = excelPath
		self.caseList = []
		self.resultList = [{'ID': 1, 'ActualResult': 'Beijing', 'Result': 'Pass', 'ResTime': '3'}, {'ID': 2, 'ActualResult': 'Failure', 'Result': 'Fail', 'ResTime': '6'}]

	def get_case_list(self):
		exceldata = xlrd.open_workbook(self.excelFile)
		sheet = exceldata.sheet_by_index(0)
		ID = sheet.col_values(0, 1)
		CityName = sheet.col_values(2, 1)
		CountryName = sheet.col_values(3, 1)
		ExpResult = sheet.col_values(4, 1)
		for i in range(len(ID)):
			temp = {}
			temp['ID'] = ID[i]
			temp['CityName'] = CityName[i]
			temp['CountryName'] = CountryName[i]
			temp['ExpResult'] = ExpResult[i]
			self.caseList.append(temp)
		print(self.caseList)
		return self.caseList

	def write_case_result(self):
		exceldata = load_workbook(self.excelFile)
		sheet = exceldata.get_sheet_by_name('Sheet1')
		for result in self.resultList:
			print(result)
			for i in range(len(self.resultList)):
				print(i)
				if result['ID'] == sheet.cell(row=i+2, column=1).value:
					print(i)
					sheet.cell(row=i+2, column=6).value = result['ActualResult']
					sheet.cell(row=i+2, column=7).value = result['Result']
					sheet.cell(row=i+2, column=8).value = result['ResTime']
		exceldata.save(self.excelFile)


excel = ExcelData(excelPath)
excel.get_case_list()
excel.write_case_result()


class GetWeather(object):
	def __init__(self, requestURL, requestBody, requestHeaders):
		self.requestURL = requestURL
		self.requestBody = requestBody
		self.requestHeaders = requestHeaders
		self.requestResult = {}

	def get_weather(self):
		starttime = time.time()
		tmp = urllib.request.
